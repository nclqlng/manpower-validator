import io
import socket
from typing import Optional
from pathlib import Path  
import textwrap

import csv
import math
from urllib.error import HTTPError, URLError
from urllib.request import urlopen

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from pandas.errors import ParserError

try:
    import msoffcrypto  # type: ignore[import-not-found]
except ImportError:
    msoffcrypto = None

try:
    import altair as alt  # type: ignore[import-not-found]
except ImportError:
    alt = None


def _altair_enabled() -> bool:
    return alt is not None


_ALTAIR_THEME_READY = False


def _altair_theme() -> None:
    global _ALTAIR_THEME_READY
    if not _altair_enabled():
        return
    if _ALTAIR_THEME_READY:
        return

    # Use the requested Sun Life palette (from CSS tokens):
    # --sunlife-gold, --sunlife-gold-light, --sunlife-navy-dark,
    # --sunlife-orange, --sunlife-teal, --sunlife-success, --sunlife-warning
    palette = [
        "#f5b400",  # sunlife gold
        "#2d4a7c",  # sunlife navy light (replaces 2nd yellow for contrast)
        "#0f1f38",  # sunlife navy dark
        "#f97316",  # sunlife orange
        "#2dd4bf",  # sunlife teal
        "#22c55e",  # sunlife success
        "#f59e0b",  # sunlife warning
        "#ef4444",  # error (still useful for negative/alerts)
    ]

    # Keep charts clean and "product" looking (lighter grid, rounded feel).
    alt.themes.register(
        "sl_modern",
        lambda: {
            "config": {
                "background": "#ffffff",
                "view": {"stroke": "transparent"},
                "range": {"category": palette},
                "axis": {
                    "labelColor": "#475569",
                    "titleColor": "#0F172A",
                    "gridColor": "rgba(15, 23, 42, 0.08)",
                    "domainColor": "rgba(15, 23, 42, 0.18)",
                    "tickColor": "rgba(15, 23, 42, 0.12)",
                    "labelFontSize": 11,
                    "titleFontSize": 12,
                    "labelAngle": 0,
                },
                "axisX": {
                    "labelAngle": 0,
                    "labelOverlap": True,
                    "labelPadding": 10,
                },
                "legend": {
                    "labelColor": "#334155",
                    "titleColor": "#0F172A",
                    "labelFontSize": 11,
                    "titleFontSize": 12,
                    "symbolType": "circle",
                    "symbolSize": 110,
                    "orient": "bottom",
                },
            }
        },
    )
    alt.themes.enable("sl_modern")
    _ALTAIR_THEME_READY = True


def _format_kpi_long(monthly_kpi: pd.DataFrame) -> pd.DataFrame:
    # Period Month + AC/NSC/Lives -> long form for Altair
    out = monthly_kpi.copy()
    out["Period Month"] = out["Period Month"].astype(str)
    long_df = out.melt(
        id_vars=["Period Month"],
        value_vars=["AC", "NSC", "Lives"],
        var_name="Metric",
        value_name="Value",
    )
    return long_df


def _render_monthly_momentum_chart(monthly_kpi: pd.DataFrame) -> None:
    if not _altair_enabled() or monthly_kpi.empty:
        st.line_chart(monthly_kpi.set_index("Period Month")[["AC", "NSC", "Lives"]], height=320)
        return

    _altair_theme()
    long_df = _format_kpi_long(monthly_kpi)

    # Keep Unknown last (if present) by pushing it to a high sort key.
    months = [m for m in monthly_kpi["Period Month"].astype(str).tolist() if m != "Unknown"]
    sort_order = months + (["Unknown"] if "Unknown" in set(long_df["Period Month"]) else [])

    base = (
        alt.Chart(long_df)
        .encode(
            x=alt.X("Period Month:N", sort=sort_order, title=None, axis=alt.Axis(labelAngle=0)),
            y=alt.Y("Value:Q", title=None),
            color=alt.Color("Metric:N", title=None),
            tooltip=[
                alt.Tooltip("Period Month:N", title="Month"),
                alt.Tooltip("Metric:N"),
                alt.Tooltip("Value:Q", format=",.2f"),
            ],
        )
        .properties(height=320)
    )

    lines = base.mark_line(strokeWidth=3, opacity=0.9)
    points = base.mark_point(size=70, filled=True, opacity=0.95)

    st.altair_chart((lines + points).interactive(), use_container_width=True)


def _render_classification_composition_chart(classification_summary: pd.DataFrame) -> None:
    if not _altair_enabled() or classification_summary.empty:
        st.bar_chart(
            classification_summary.set_index("Classification")[["AC", "NSC", "Lives"]],
            height=330,
        )
        return

    _altair_theme()
    long_df = classification_summary.melt(
        id_vars=["Classification"], value_vars=["AC", "NSC", "Lives"], var_name="Metric", value_name="Value"
    )

    chart = (
        alt.Chart(long_df)
        .mark_bar(cornerRadiusTopLeft=6, cornerRadiusTopRight=6)
        .encode(
            x=alt.X("Classification:N", title=None),
            y=alt.Y("Value:Q", title=None),
            xOffset=alt.XOffset("Metric:N"),
            color=alt.Color("Metric:N", title=None),
            tooltip=[
                alt.Tooltip("Classification:N"),
                alt.Tooltip("Metric:N"),
                alt.Tooltip("Value:Q", format=",.2f"),
            ],
        )
        .properties(height=330)
    )

    st.altair_chart(chart, use_container_width=True)


def _render_top_advisors_chart(top_advisors: pd.DataFrame, ranking_metric: str) -> None:
    if not _altair_enabled() or top_advisors.empty:
        st.bar_chart(top_advisors.set_index("Advisor")[["AC", "NSC", "Lives"]], height=330)
        return

    _altair_theme()
    metric = ranking_metric
    df = top_advisors.copy()
    df["Advisor"] = df["Advisor"].astype(str)
    df = df.sort_values(metric, ascending=False).reset_index(drop=True)
    df["Rank"] = df.index + 1
    df["Rank Label"] = "Rank " + df["Rank"].astype(str)

    rank_palette_seed = [
        "#f5b400",  # Rank 1 - sunlife gold
        "#2d4a7c",  # Rank 2 - sunlife navy light
        "#0f1f38",  # Rank 3 - sunlife navy dark
        "#f97316",  # Rank 4 - sunlife orange
        "#2dd4bf",  # Rank 5 - sunlife teal
        "#22c55e",  # Rank 6 - success
        "#f59e0b",  # Rank 7 - warning
    ]
    rank_colors = [rank_palette_seed[i % len(rank_palette_seed)] for i in range(len(df))]
    rank_domain = df["Rank Label"].tolist()

    chart = (
        alt.Chart(df)
        .mark_bar(cornerRadiusTopLeft=7, cornerRadiusTopRight=7)
        .encode(
            y=alt.Y("Advisor:N", sort="-x", title=None),
            x=alt.X(f"{metric}:Q", title=None),
            color=alt.Color(
                "Rank Label:N",
                title=None,
                legend=None,
                scale=alt.Scale(domain=rank_domain, range=rank_colors),
            ),
            tooltip=[
                alt.Tooltip("Rank:Q"),
                alt.Tooltip("Advisor:N"),
                alt.Tooltip("AC:Q", format=",.2f"),
                alt.Tooltip("NSC:Q", format=",.2f"),
                alt.Tooltip("Lives:Q", format=",.0f"),
            ],
        )
        .properties(height=330)
    )

    st.altair_chart(chart, use_container_width=True)


def _render_monthly_ac_by_classification(monthly_summary: pd.DataFrame, ordered_months: list[str]) -> None:
    if not _altair_enabled() or monthly_summary.empty:
        pivot = (
            monthly_summary.pivot(index="Period Month", columns="Classification", values="AC")
            .fillna(0)
            .reindex([m for m in ordered_months if m in set(monthly_summary["Period Month"])])
        )
        st.bar_chart(pivot, height=320)
        return

    _altair_theme()
    df = monthly_summary.copy()
    df["Period Month"] = df["Period Month"].astype(str)

    chart = (
        alt.Chart(df)
        .mark_bar(cornerRadiusTopLeft=5, cornerRadiusTopRight=5)
        .encode(
            x=alt.X("Period Month:N", sort=ordered_months, title=None, axis=alt.Axis(labelAngle=0)),
            y=alt.Y("AC:Q", title=None, stack=True),
            color=alt.Color("Classification:N", title=None),
            tooltip=[
                alt.Tooltip("Period Month:N", title="Month"),
                alt.Tooltip("Classification:N"),
                alt.Tooltip("AC:Q", format=",.2f"),
            ],
        )
        .properties(height=320)
    )

    st.altair_chart(chart, use_container_width=True)



def apply_theme() -> None:
    """Load and apply CSS theme from external file"""
    css_path = Path(__file__).parent / "app_styles.css"
    if css_path.exists():
        css_text = css_path.read_text(encoding="utf-8")
        st.markdown(f"<style>{css_text}</style>", unsafe_allow_html=True)

def render_hero() -> None:
    st.markdown(
        textwrap.dedent(
            """
            <div class="sl-hero">
              <div class="sl-hero-top">
                <div class="sl-hero-badge">Sun Life • Manpower Analytics</div>
                <div class="sl-hero-actions">
                  <span class="sl-pill">
                    <svg viewBox="0 0 24 24" aria-hidden="true"><path d="M12 3v12"/><path d="m7 10 5 5 5-5"/><path d="M5 21h14"/></svg>
                    Upload
                  </span>
                  <span class="sl-pill">
                    <svg viewBox="0 0 24 24" aria-hidden="true"><path d="M3 6h18"/><path d="M7 6v14"/><path d="M3 10h4"/><path d="M3 14h4"/><path d="M3 18h4"/></svg>
                    Map
                  </span>
                  <span class="sl-pill">
                    <svg viewBox="0 0 24 24" aria-hidden="true"><path d="M3 3v18h18"/><path d="m7 14 3-3 3 2 5-6"/></svg>
                    Analyze
                  </span>
                  <span class="sl-pill sl-pill-ghost">
                    <svg viewBox="0 0 24 24" aria-hidden="true"><path d="M12 14v7"/><path d="M8 17l4 4 4-4"/><path d="M20 13a4 4 0 0 0-3-3.87"/><path d="M4 13a4 4 0 0 1 3-3.87"/><path d="M7 9a5 5 0 0 1 10 0"/></svg>
                    Export
                  </span>
                </div>
              </div>

              <div class="sl-hero-title">Manpower Validation</div>
              <div class="sl-hero-sub">
                Turn monthly &amp; quarterly production data into a clean story: performance,
                validation status, drivers, trends, and export-ready reports.
              </div>

              <div class="sl-hero-meta">
                <div class="sl-hero-meta-item">
                  <div class="sl-hero-meta-k">Best for</div>
                  <div class="sl-hero-meta-v">Dashboard / Settled / Submitted sheets</div>
                </div>
                <div class="sl-hero-meta-item">
                  <div class="sl-hero-meta-k">Output</div>
                  <div class="sl-hero-meta-v">Excel report + corrected Dashboard (XLSX)</div>
                </div>
                <div class="sl-hero-meta-item">
                  <div class="sl-hero-meta-k">Includes</div>
                  <div class="sl-hero-meta-v">AC / NSC / Lives + Validation snapshot</div>
                </div>
              </div>
            </div>
            """
        ),
        unsafe_allow_html=True,
    )


def render_section(title: str, subtitle: str) -> None:
    st.markdown(
        f"""
        <div class="sl-section">
            <div class="sl-section-title">{title}</div>
            <div class="sl-section-sub">{subtitle}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_chart_card(title: str, render_fn) -> None:
    with st.container(border=True):
        st.markdown(f'<div class="sl-chart-title">{title}</div>', unsafe_allow_html=True)
        render_fn()


st.set_page_config(page_title="Manpower Validation System", layout="wide")
apply_theme()
render_hero()


def normalize_number(series: pd.Series) -> pd.Series:
    cleaned = (
        series.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace(" ", "", regex=False)
        .replace({"": "0", "nan": "0", "None": "0"})
    )
    return pd.to_numeric(cleaned, errors="coerce").fillna(0)


# ---- 2026 Advisor Code -> Classification mapping (Google Sheets) ----
GSHEET_2026_BASE_URL = (
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vRGyFGolZgI4lBcvNQoywRotN6i8oihoONR53pdD2RQomtVl1RwVieG2M2Vn-zM5A"
)
GSHEET_2026_GID = "2022092860"


@st.cache_data(ttl=24 * 3600)
def fetch_2026_advisor_code_to_classification() -> tuple[dict[str, str], int]:
    """
    Fetches the published 2026 tab as CSV, then builds:
      Advisor Code (string) -> Classification (UPPER string)

    Note: the CSV includes extra header rows; data rows start with a numeric index.
    """

    url = f"{GSHEET_2026_BASE_URL}/pub?gid={GSHEET_2026_GID}&single=true&output=csv"
    raw = ""
    for timeout_seconds in (15, 30):
        try:
            raw = urlopen(url, timeout=timeout_seconds).read().decode("utf-8", errors="replace")
            break
        except (TimeoutError, socket.timeout, URLError, HTTPError, OSError):
            continue

    if not raw:
        return {}, 0

    mapping: dict[str, str] = {}
    try:
        for row in csv.reader(io.StringIO(raw)):
            if not row or len(row) < 5:
                continue
            row_idx = row[0].strip() if row[0] is not None else ""
            if not row_idx.isdigit():
                continue

            # Based on the published CSV structure:
            #   0: row index, 1: advisor name, 2: advisor code, 3: coding date, 4: classification
            code = (row[2] or "").strip()
            cls = (row[4] or "").strip().upper()
            if not code or not cls:
                continue
            mapping[code] = cls  # later rows override earlier ones
    except (csv.Error, OSError):
        return {}, 0

    target_code_len = max((len(k) for k in mapping.keys()), default=0)
    return mapping, target_code_len


def normalize_advisor_code_value(value: object, target_len: int) -> Optional[str]:
    if value is None:
        return None

    code_str: str
    if isinstance(value, str):
        code_str = value.strip().replace("\u00a0", " ")
    elif isinstance(value, bool):
        return None
    elif isinstance(value, int):
        code_str = str(value)
    elif isinstance(value, float):
        if math.isnan(value):
            return None
        code_str = str(int(value)) if value.is_integer() else str(value).split(".")[0]
    else:
        code_str = str(value).strip()

    if not code_str or code_str.lower() in {"nan", "none"}:
        return None

    # Common case: Excel stores codes as numeric and pandas/openpyxl may represent them as X.0
    if code_str.endswith(".0"):
        maybe_int = code_str[:-2]
        if maybe_int.isdigit():
            code_str = maybe_int

    if target_len and code_str.isdigit() and len(code_str) < target_len:
        code_str = code_str.zfill(target_len)

    return code_str


def normalize_advisor_code_series(series: pd.Series, target_len: int) -> pd.Series:
    s = series.astype(str).str.strip().replace({"nan": "", "None": ""})
    s = s.str.replace(r"\.0$", "", regex=True)
    if target_len:
        mask = s.str.fullmatch(r"\d+")
        s.loc[mask] = s.loc[mask].str.zfill(target_len)
    return s.replace({"": pd.NA})


def correct_classifications_in_workbook(
    file_bytes: bytes,
    sheet_name: str,
    header_row: int,
    advisor_code_header: str,
    classification_header: str,
    code_to_class: dict[str, str],
    target_code_len: int,
    excel_password: Optional[str],
) -> tuple[bytes, int]:
    """
    Updates only the Classification cell value (not formatting) in-place using openpyxl.
    """

    workbook_bytes, _ = maybe_decrypt_excel_bytes(file_bytes, excel_password or "")
    wb = load_workbook(io.BytesIO(workbook_bytes))
    if sheet_name not in wb.sheetnames:
        return file_bytes, 0

    ws = wb[sheet_name]

    header_norm_to_col: dict[str, int] = {}
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        key = str(cell.value).strip().upper()
        if key:
            header_norm_to_col.setdefault(key, cell.column)

    advisor_key = str(advisor_code_header).strip().upper()
    class_key = str(classification_header).strip().upper()
    if advisor_key not in header_norm_to_col or class_key not in header_norm_to_col:
        return file_bytes, 0

    advisor_col_idx = header_norm_to_col[advisor_key]
    class_col_idx = header_norm_to_col[class_key]

    updated = 0
    for r in range(header_row + 1, ws.max_row + 1):
        code_val = ws.cell(row=r, column=advisor_col_idx).value
        code_norm = normalize_advisor_code_value(code_val, target_code_len)
        if not code_norm:
            continue

        new_cls = code_to_class.get(code_norm)
        if not new_cls:
            continue

        cls_cell = ws.cell(row=r, column=class_col_idx)
        current = cls_cell.value
        current_norm = str(current).strip().upper() if current is not None else ""
        if current_norm != new_cls:
            cls_cell.value = new_cls
            updated += 1

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), updated


def build_period_columns(
    df: pd.DataFrame,
    date_col: Optional[str],
    month_col: Optional[str],
    year_col: Optional[str],
) -> pd.DataFrame:
    out = df.copy()
    parsed_date = pd.Series(pd.NaT, index=out.index, dtype="datetime64[ns]")

    if date_col:
        date_text = out[date_col].astype(str).str.strip()
        # Primary format: "April 1, 2026"
        parsed_date = pd.to_datetime(date_text, format="%B %d, %Y", errors="coerce")
        # Fallback for other valid date shapes.
        parsed_date = parsed_date.fillna(pd.to_datetime(date_text, errors="coerce"))
    elif month_col and year_col:
        month_raw = out[month_col].astype(str).str.strip()
        year_raw = out[year_col].astype(str).str.extract(r"(\d{4})")[0]
        month_num = pd.to_numeric(month_raw, errors="coerce")
        month_from_name = pd.to_datetime(month_raw, format="%B", errors="coerce").dt.month
        month_final = month_num.fillna(month_from_name)
        composed = year_raw.fillna("") + "-" + month_final.fillna(1).astype(int).astype(str) + "-01"
        parsed_date = pd.to_datetime(composed, errors="coerce")

    out["Period Month"] = parsed_date.dt.to_period("M").astype(str)
    out["Period Quarter"] = parsed_date.dt.to_period("Q").astype(str)
    out["Period Year"] = parsed_date.dt.year.astype("Int64").astype(str)
    out["Period Month"] = out["Period Month"].replace("NaT", "Unknown")
    out["Period Quarter"] = out["Period Quarter"].replace("NaT", "Unknown")
    out["Period Year"] = out["Period Year"].replace("<NA>", "Unknown")
    return out


def guess_column(columns: list[str], preferred_keywords: list[str]) -> str:
    lowered = [(col, col.lower()) for col in columns]
    for keyword in preferred_keywords:
        for original, lowered_col in lowered:
            if keyword in lowered_col:
                return original
    return columns[0]


def series_is_mostly_numeric(series: pd.Series) -> bool:
    cleaned = series.astype(str).str.strip()
    cleaned = cleaned[~cleaned.isin(["", "nan", "None"])]
    if cleaned.empty:
        return False
    numeric_ratio = pd.to_numeric(cleaned, errors="coerce").notna().mean()
    return numeric_ratio >= 0.8


def sort_period_labels(labels: list[str], freq: str) -> list[str]:
    known = [label for label in labels if label != "Unknown"]
    ordered = sorted(known, key=lambda x: pd.Period(x, freq=freq))
    if "Unknown" in labels:
        ordered.append("Unknown")
    return ordered


def to_excel_bytes(sheets: dict[str, pd.DataFrame]) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, data in sheets.items():
            sheet_title = sheet_name[:31]
            data.to_excel(writer, index=False, sheet_name=sheet_title)
            if sheet_title == "Validation Results":
                ws = writer.book[sheet_title]
                _add_validation_results_excel_logic(ws)
    return buffer.getvalue()


def _add_validation_results_excel_logic(ws) -> None:
    headers = [cell.value for cell in ws[1]]
    header_to_col = {str(header): idx + 1 for idx, header in enumerate(headers) if header is not None}
    required_headers = [
        "Classification",
        "Tenure Raw",
        "Coding Quarter",
        "AC",
        "NSC",
        "JFW Done",
        "START Done",
        "Pillars Done",
        "VUL Advance Done",
        "Mandatory Training Done",
        "Validation Requirement",
        "Validation Status",
        "Validation Reason",
        "VNA Eligible",
        "VMP Eligible",
        "SM Appointment Eligible",
    ]
    if not all(header in header_to_col for header in required_headers):
        return

    last_row = ws.max_row
    if last_row < 2:
        return

    jfw_col = get_column_letter(header_to_col["JFW Done"])
    start_col = get_column_letter(header_to_col["START Done"])
    pillars_col = get_column_letter(header_to_col["Pillars Done"])
    vul_col = get_column_letter(header_to_col["VUL Advance Done"])
    cls_col = get_column_letter(header_to_col["Classification"])
    tenure_col = get_column_letter(header_to_col["Tenure Raw"])
    coding_q_col = get_column_letter(header_to_col["Coding Quarter"])
    ac_col = get_column_letter(header_to_col["AC"])
    nsc_col = get_column_letter(header_to_col["NSC"])
    mandatory_col = get_column_letter(header_to_col["Mandatory Training Done"])
    requirement_col = get_column_letter(header_to_col["Validation Requirement"])
    status_col = get_column_letter(header_to_col["Validation Status"])
    reason_col = get_column_letter(header_to_col["Validation Reason"])
    vna_col = get_column_letter(header_to_col["VNA Eligible"])
    vmp_col = get_column_letter(header_to_col["VMP Eligible"])
    sm_col = get_column_letter(header_to_col["SM Appointment Eligible"])

    # Add TRUE/FALSE dropdowns for training columns in downloaded Excel.
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    ws.add_data_validation(dv)
    for col in (jfw_col, start_col, pillars_col, vul_col, mandatory_col):
        dv.add(f"{col}2:{col}{last_row}")

    for row in range(2, last_row + 1):
        cls = f"UPPER(TRIM({cls_col}{row}&\"\"))"
        tenure = f"LOWER({tenure_col}{row}&\"\")"
        coding_q = f"{coding_q_col}{row}"
        ac_value = f"N({ac_col}{row})"
        nsc_value = f"N({nsc_col}{row})"
        jfw_true = f"OR({jfw_col}{row}=TRUE,UPPER({jfw_col}{row}&\"\")=\"TRUE\")"
        start_true = f"OR({start_col}{row}=TRUE,UPPER({start_col}{row}&\"\")=\"TRUE\")"
        pillars_true = f"OR({pillars_col}{row}=TRUE,UPPER({pillars_col}{row}&\"\")=\"TRUE\")"
        vul_true = f"OR({vul_col}{row}=TRUE,UPPER({vul_col}{row}&\"\")=\"TRUE\")"
        mandatory_true = f"OR({mandatory_col}{row}=TRUE,UPPER({mandatory_col}{row}&\"\")=\"TRUE\")"
        training_bundle_ok = f"AND({jfw_true},{start_true},{pillars_true})"
        is_external_mc = f"AND({cls}=\"MC\",ISNUMBER(SEARCH(\"external\",{tenure})))"
        is_rookie = (
            f"OR(ISNUMBER(SEARCH(\"year 0\",{tenure})),"
            f"ISNUMBER(SEARCH(\"rookie\",{tenure})),"
            f"ISNUMBER(SEARCH(\"external\",{tenure})))"
        )
        both_45k = f"AND({ac_value}>=45000,{nsc_value}>=45000)"
        both_90k = f"AND({ac_value}>=90000,{nsc_value}>=90000)"
        both_135k = f"AND({ac_value}>=135000,{nsc_value}>=135000)"
        both_180k = f"AND({ac_value}>=180000,{nsc_value}>=180000)"
        rookie_threshold = f"IF({coding_q}=4,15000,45000)"
        both_rookie_threshold = f"AND({ac_value}>={rookie_threshold},{nsc_value}>={rookie_threshold})"
        passed_expr = (
            f"IF({is_external_mc},AND({both_45k},{training_bundle_ok}),"
            f"IF({cls}=\"A\",AND({both_rookie_threshold},{training_bundle_ok}),"
            f"IF({cls}=\"B\",AND({both_90k},{vul_true}),"
            f"IF({cls}=\"C\",{both_135k},"
            f"IF(OR({cls}=\"D\",{cls}=\"E\",{cls}=\"MC\",{cls}=\"F\"),{both_180k},FALSE)))))"
        )

        ws[f"{requirement_col}{row}"] = (
            f"=IF({is_external_mc},\"External MC: AC>=45K and NSC>=45K + JFW + START + 4 Pillars\","
            f"IF({cls}=\"A\",IF({coding_q}=4,\"Rookie A: AC>=15,000 and NSC>=15,000 + JFW + START + 4 Pillars\","
            f"\"Rookie A: AC>=45,000 and NSC>=45,000 + JFW + START + 4 Pillars\"),"
            f"IF({cls}=\"B\",\"Tenured B: AC>=90K and NSC>=90K + VUL Advance\","
            f"IF({cls}=\"C\",\"Tenured C: AC>=135K and NSC>=135K\","
            f"IF(OR({cls}=\"D\",{cls}=\"E\",{cls}=\"MC\"),\"Tenured D/E/MC: AC>=180K and NSC>=180K\","
            f"IF({cls}=\"F\",\"F advisor: optional, counted as VMP if AC>=180K and NSC>=180K\",\"No mapped rule\"))))))"
        )
        ws[f"{status_col}{row}"] = f"=IF({passed_expr},\"Pass\",\"Fail\")"
        ws[f"{reason_col}{row}"] = (
            f"=IF({passed_expr},"
            f"IF({cls}=\"F\",\"Counts as VMP via AC>=180K and NSC>=180K.\",\"Meets requirement.\"),"
            f"IF({is_external_mc},\"Needs AC>=45K, NSC>=45K, and complete JFW/START/4 Pillars.\","
            f"IF({cls}=\"A\",\"Needs AC/NSC threshold (both) and complete JFW/START/4 Pillars.\","
            f"IF({cls}=\"B\",\"Needs AC>=90K, NSC>=90K, and VUL Advance completion.\","
            f"IF({cls}=\"C\",\"Needs AC>=135K and NSC>=135K.\","
            f"IF(OR({cls}=\"D\",{cls}=\"E\",{cls}=\"MC\"),\"Needs AC>=180K and NSC>=180K.\","
            f"IF({cls}=\"F\",\"Below optional 180K VMP threshold for AC and NSC.\","
            f"\"Classification/tenure combination not mapped to a rule.\")))))))"
        )
        ws[f"{vna_col}{row}"] = f"=IF(AND({is_rookie},{passed_expr}),\"Yes\",\"No\")"
        ws[f"{vmp_col}{row}"] = (
            f"=IF(OR({passed_expr},AND({is_rookie},{both_90k},{training_bundle_ok})),\"Yes\",\"No\")"
        )
        ws[f"{sm_col}{row}"] = f"=IF(AND({both_180k},{mandatory_true}),\"Yes\",\"No\")"


def performance_status(actual: float, target: float) -> str:
    if target <= 0:
        return "No target set"
    ratio = actual / target
    if ratio >= 1:
        return "On track (target achieved)"
    if ratio >= 0.8:
        return "Needs attention (close to target)"
    return "Off track (below target)"


def format_compact(value: float) -> str:
    abs_val = abs(value)
    if abs_val >= 1_000_000_000:
        return f"{value / 1_000_000_000:.2f}B"
    if abs_val >= 1_000_000:
        return f"{value / 1_000_000:.2f}M"
    if abs_val >= 1_000:
        return f"{value / 1_000:.1f}K"
    return f"{value:,.2f}" if value % 1 else f"{value:,.0f}"


def normalize_flag(series: pd.Series) -> pd.Series:
    truthy = {"1", "y", "yes", "true", "completed", "done", "pass", "passed"}
    return series.astype(str).str.strip().str.lower().isin(truthy)


def evaluate_sunlife_validation(row: pd.Series) -> dict[str, object]:
    cls = str(row.get("Classification", "")).strip().upper()
    tenure_text = str(row.get("Tenure Raw", "")).strip().lower()
    ac = float(row.get("AC", 0) or 0)
    nsc = float(row.get("NSC", 0) or 0)
    coding_q = int(row.get("Coding Quarter", 0) or 0)
    jfw = bool(row.get("JFW Done", False))
    start = bool(row.get("START Done", False))
    pillars = bool(row.get("Pillars Done", False))
    vul = bool(row.get("VUL Advance Done", False))
    mandatory_training = bool(row.get("Mandatory Training Done", False))

    is_rookie = ("year 0" in tenure_text) or ("rookie" in tenure_text) or ("external" in tenure_text)
    is_external_mc = cls == "MC" and "external" in tenure_text
    training_bundle_ok = jfw and start and pillars
    meets = lambda threshold: ac >= threshold and nsc >= threshold

    requirement = "No mapped rule"
    passed = False
    reason = "Classification/tenure combination not mapped to a rule."

    if is_external_mc:
        requirement = "External MC: AC>=45K and NSC>=45K + JFW + START + 4 Pillars"
        passed = meets(45_000) and training_bundle_ok
        reason = "Meets requirement." if passed else "Needs AC>=45K, NSC>=45K, and complete JFW/START/4 Pillars."
    elif cls == "A":
        threshold = 15_000 if coding_q == 4 else 45_000
        requirement = f"Rookie A: AC>={threshold:,.0f} and NSC>={threshold:,.0f} + JFW + START + 4 Pillars"
        passed = meets(threshold) and training_bundle_ok
        reason = "Meets requirement." if passed else "Needs AC/NSC threshold (both) and complete JFW/START/4 Pillars."
    elif cls == "B":
        requirement = "Tenured B: AC>=90K and NSC>=90K + VUL Advance"
        passed = meets(90_000) and vul
        reason = "Meets requirement." if passed else "Needs AC>=90K, NSC>=90K, and VUL Advance completion."
    elif cls == "C":
        requirement = "Tenured C: AC>=135K and NSC>=135K"
        passed = meets(135_000)
        reason = "Meets requirement." if passed else "Needs AC>=135K and NSC>=135K."
    elif cls in {"D", "E", "MC"}:
        requirement = "Tenured D/E/MC: AC>=180K and NSC>=180K"
        passed = meets(180_000)
        reason = "Meets requirement." if passed else "Needs AC>=180K and NSC>=180K."
    elif cls == "F":
        requirement = "F advisor: optional, counted as VMP if AC>=180K and NSC>=180K"
        passed = meets(180_000)
        reason = "Counts as VMP via AC>=180K and NSC>=180K." if passed else "Below optional 180K VMP threshold for AC and NSC."

    vna = is_rookie and passed
    rookie_vmp = is_rookie and meets(90_000) and training_bundle_ok
    sm_eligible = meets(180_000) and mandatory_training
    vmp = passed or rookie_vmp

    return {
        "Validation Requirement": requirement,
        "Validation Status": "Pass" if passed else "Fail",
        "Validation Reason": reason,
        "VNA Eligible": "Yes" if vna else "No",
        "VMP Eligible": "Yes" if vmp else "No",
        "SM Appointment Eligible": "Yes" if sm_eligible else "No",
    }


TRAINING_COLUMNS = ["JFW Done", "START Done", "Pillars Done", "VUL Advance Done"]
EDITABLE_VALIDATION_COLUMNS = TRAINING_COLUMNS + ["Mandatory Training Done"]


def required_training_by_row(row: pd.Series) -> dict[str, bool]:
    cls = str(row.get("Classification", "")).strip().upper()
    tenure_text = str(row.get("Tenure Raw", "")).strip().lower()
    is_external_mc = cls == "MC" and "external" in tenure_text
    requires_bundle = cls == "A" or is_external_mc
    requires_vul = cls == "B"
    return {
        "JFW Done": requires_bundle,
        "START Done": requires_bundle,
        "Pillars Done": requires_bundle,
        "VUL Advance Done": requires_vul,
    }


def enforce_training_requirements(df: pd.DataFrame) -> pd.DataFrame:
    adjusted = df.copy()
    for idx, row in adjusted.iterrows():
        required = required_training_by_row(row)
        for col in TRAINING_COLUMNS:
            if not required[col]:
                adjusted.at[idx, col] = False
    return adjusted


uploaded_files = st.file_uploader(
    "Upload Excel file(s)", type=["xlsx", "xlsm", "xls"], accept_multiple_files=True
)

# Optional password for encrypted Excel files.
# Note: xlrd (for legacy `.xls`) is the primary engine that supports `password` via pandas.
excel_password = st.text_input(
    "Excel password (optional, used for password-protected Excel files)",
    type="password",
    value="",
)

# Detect whether an uploaded Excel file is truly legacy `.xls` (OLE2) or
# newer `.xlsx/.xlsm` (ZIP) based on the raw file signature.
def detect_excel_kind(file_bytes: bytes) -> str:
    # ZIP magic for OOXML: XLSX/XLSM
    if file_bytes.startswith(b"PK\x03\x04"):
        return "xlsx"
    # OLE2 compound document magic for XLS
    if file_bytes.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"):
        return "xls"
    return "unknown"


def is_office_file_encrypted(file_bytes: bytes) -> bool:
    """
    Best-effort check for password-protected Office files.
    Encrypted modern Excel workbooks commonly use an OLE container, so they can
    look like legacy `.xls` files from their first bytes alone.
    """
    if msoffcrypto is None:
        return False

    try:
        office_file = msoffcrypto.OfficeFile(io.BytesIO(file_bytes))
        return bool(office_file.is_encrypted())
    except Exception:
        return False


def maybe_decrypt_excel_bytes(file_bytes: bytes, password: str) -> tuple[bytes, bool]:
    """
    Decrypt password-protected Office files (including encrypted `.xlsx`) when possible.
    Returns `(bytes_to_read, was_decrypted)`.
    """
    if not password:
        return file_bytes, False
    if msoffcrypto is None:
        return file_bytes, False

    try:
        office_file = msoffcrypto.OfficeFile(io.BytesIO(file_bytes))
    except Exception:
        return file_bytes, False
    if not office_file.is_encrypted():
        return file_bytes, False

    office_file.load_key(password=password)
    decrypted = io.BytesIO()
    office_file.decrypt(decrypted)
    return decrypted.getvalue(), True

# Streamlit re-runs the script on every widget change. File uploads are usually
# preserved by the widget, but storing bytes in session_state makes this
# behavior explicit and avoids losing them on re-runs.
if uploaded_files:
    file_payloads = [(f.name, f.read()) for f in uploaded_files]
    st.session_state["uploaded_file_payloads"] = file_payloads
else:
    file_payloads = st.session_state.get("uploaded_file_payloads")

if not file_payloads:
    st.info("📂 Upload one or more files to start.")
    st.stop()

# Pick engine based on file bytes signature, but also consider filename extension.
# Signature detection can be wrong for mislabelled uploads, truncated files, or proxies.
first_file_name = file_payloads[0][0] or ""
first_bytes = file_payloads[0][1]
first_raw_kind = detect_excel_kind(first_bytes)
first_is_encrypted = is_office_file_encrypted(first_bytes)
first_ext = first_file_name.rsplit(".", 1)[-1].lower() if "." in first_file_name else ""

if first_is_encrypted and not excel_password:
    st.error(
        "The first uploaded workbook appears to be password-protected.\n\n"
        f"Filename: `{first_file_name}`\n"
        f"Signature detected kind: `{first_raw_kind}`\n"
        f"Filename extension: `{first_ext}`\n\n"
        "Please enter the workbook password and try again. "
        "Encrypted modern `.xlsx` files often have an OLE signature, so they can "
        "look like legacy `.xls` files until they are decrypted."
    )
    st.stop()

try:
    prepared_first_bytes, first_was_decrypted = maybe_decrypt_excel_bytes(first_bytes, excel_password)
except Exception as e:
    st.error(f"Unable to decrypt `{first_file_name}`. Check the Excel password and try again.\n\nDetails: {e}")
    st.stop()
first_kind = detect_excel_kind(prepared_first_bytes)


def engine_kwargs_for(*, extra_kwargs: Optional[dict] = None) -> dict:
    base_kwargs: dict = {}
    if extra_kwargs:
        base_kwargs.update(extra_kwargs)
    return base_kwargs


engine_attempts: list[tuple[str, dict]] = []
attempt_dedup_keys: set[tuple[str, bool]] = set()


def add_attempt(engine_name: str, *, extra_kwargs: Optional[dict] = None) -> None:
    extra_kwargs = extra_kwargs or {}
    ignore_corruption = bool(extra_kwargs.get("ignore_workbook_corruption"))
    key = (engine_name, ignore_corruption)
    if key in attempt_dedup_keys:
        return
    attempt_dedup_keys.add(key)
    engine_attempts.append((engine_name, engine_kwargs_for(extra_kwargs=extra_kwargs)))


ext_preferred_engine: Optional[str] = None
if first_ext in {"xlsx", "xlsm"}:
    ext_preferred_engine = "openpyxl"
elif first_ext == "xls":
    ext_preferred_engine = "xlrd"

signature_preferred_engine: Optional[str] = None
if first_kind == "xlsx":
    signature_preferred_engine = "openpyxl"
elif first_kind == "xls":
    signature_preferred_engine = "xlrd"

if ext_preferred_engine:
    add_attempt(ext_preferred_engine)
if signature_preferred_engine and signature_preferred_engine != ext_preferred_engine:
    add_attempt(signature_preferred_engine)

# Always try the other engine as a fallback.
if ext_preferred_engine == "openpyxl":
    add_attempt("xlrd")
elif ext_preferred_engine == "xlrd":
    add_attempt("openpyxl")
else:
    # Unknown extension: fall back to signature-based choice, then the other engine.
    if signature_preferred_engine == "xlrd":
        add_attempt("openpyxl")
    else:
        add_attempt("xlrd")

# Best-effort: sometimes xlrd can be coaxed past minor corruption.
if any(engine_name == "xlrd" for engine_name, _ in engine_attempts):
    add_attempt("xlrd", extra_kwargs={"ignore_workbook_corruption": True})


excel = None
attempt_errors: list[str] = []
for engine_name, engine_kwargs in engine_attempts:
    try:
        excel = pd.ExcelFile(io.BytesIO(prepared_first_bytes), engine=engine_name, engine_kwargs=engine_kwargs)
        attempt_errors = []
        break
    except Exception as e:
        attempt_errors.append(f"- {engine_name} with {engine_kwargs}: {e}")

if excel is None:
    st.error(
        "Unable to read the first uploaded Excel workbook.\n\n"
        f"Filename: `{first_file_name}`\n"
        f"Signature detected kind: `{first_kind}`\n"
        f"Filename extension: `{first_ext}`\n"
        f"Decrypted before loading: `{first_was_decrypted}`\n"
        f"First 8 bytes (hex): `{prepared_first_bytes[:8].hex()}`\n\n"
        "Engine attempts:\n"
        + "\n".join(attempt_errors[:6])
    )
    st.stop()

preferred_sheets = ["Settled Apps - Details", "Submitted Apps - Details"]
default_sheet_index = 0
for preferred_name in preferred_sheets:
    if preferred_name in excel.sheet_names:
        default_sheet_index = excel.sheet_names.index(preferred_name)
        break

sheet = st.selectbox("Sheet", excel.sheet_names, index=default_sheet_index)
header_row = st.number_input(
    "Header row number in Excel (1-based)",
    min_value=1,
    value=12,
    step=1,
    help="Set this to 12 if your actual column names start on row 12.",
)
column_range = st.text_input(
    "Column range to read (Excel style)",
    value="",
    help="Optional. Example: A:Z or A:AA. Leave blank to read all available columns.",
)
requested_range = column_range.strip()
loaded_frames = []
loaded_file_names = []
skipped_file_names = []
for file_name, file_bytes in file_payloads:
    if is_office_file_encrypted(file_bytes) and not excel_password:
        st.warning(
            f"Skipped `{file_name}` because it appears to be password-protected. "
            "Enter the Excel password to load it."
        )
        skipped_file_names.append(file_name)
        continue

    try:
        prepared_file_bytes, _ = maybe_decrypt_excel_bytes(file_bytes, excel_password)
    except Exception as e:
        st.warning(f"Skipped `{file_name}` because it could not be decrypted with the supplied password: {e}")
        skipped_file_names.append(file_name)
        continue

    kind = detect_excel_kind(prepared_file_bytes)
    file_lower = (file_name or "").lower()
    ext = file_lower.rsplit(".", 1)[-1] if "." in file_lower else ""
    try:
        engine_candidates: list[str]
        if ext in {"xlsx", "xlsm"}:
            # xlrd does not support .xlsx; keep this path strictly openpyxl.
            engine_candidates = ["openpyxl"]
        elif ext == "xls":
            engine_candidates = ["xlrd", "openpyxl"]
        elif kind == "xls":
            engine_candidates = ["xlrd"]
        elif kind == "xlsx":
            engine_candidates = ["openpyxl"]
        else:
            engine_candidates = ["openpyxl", "xlrd"]

        last_err: Optional[Exception] = None
        succeeded_engine: Optional[str] = None
        frame = None
        for candidate in engine_candidates:
            file_engine_kwargs = {}
            try:
                frame = pd.read_excel(
                    io.BytesIO(prepared_file_bytes),
                    sheet_name=sheet,
                    header=int(header_row) - 1,
                    engine=candidate,
                    engine_kwargs=file_engine_kwargs,
                )
                last_err = None
                succeeded_engine = candidate
                break
            except Exception as e:
                last_err = e
                frame = None
        if frame is None:
            raise last_err if last_err else RuntimeError("Unable to read uploaded Excel file.")

        if requested_range:
            try:
                # Re-read using the same engine that succeeded above (best-effort).
                succeeded = False
                for candidate in ([succeeded_engine] if succeeded_engine else engine_candidates):
                    file_engine_kwargs = {}
                    try:
                        frame = pd.read_excel(
                            io.BytesIO(prepared_file_bytes),
                            sheet_name=sheet,
                            header=int(header_row) - 1,
                            usecols=requested_range,
                            engine=candidate,
                            engine_kwargs=file_engine_kwargs,
                        )
                        succeeded = True
                        break
                    except Exception:
                        continue
                if not succeeded:
                    raise
            except ParserError:
                st.warning(
                    f"`{file_name}`: column range `{requested_range}` is wider than this sheet. "
                    "Loaded all available columns instead."
                )
        frame = frame.dropna(how="all")
        frame.columns = [str(c).strip() for c in frame.columns]
        frame = frame.loc[:, ~pd.Index(frame.columns).str.startswith("Unnamed:")]
        if not frame.empty:
            frame["Source File"] = file_name
            loaded_frames.append(frame)
            loaded_file_names.append(file_name)
    except ValueError:
        skipped_file_names.append(file_name)

if skipped_file_names:
    st.warning(
        f"Skipped {len(skipped_file_names)} file(s) that do not have sheet `{sheet}`: "
        + ", ".join(skipped_file_names)
    )

if not loaded_frames:
    st.warning("No valid rows were loaded from uploaded files.")
    st.stop()

raw_df = pd.concat(loaded_frames, ignore_index=True, sort=False)

if raw_df.empty:
    st.warning("Selected sheet is empty across uploaded files.")
    st.stop()

st.caption(
    f"Loaded {len(loaded_file_names)} file(s), {len(raw_df):,} rows. "
    f"Detected columns: {', '.join(raw_df.columns)}"
)
with st.expander("Loaded files summary", expanded=False):
    file_summary = (
        raw_df.groupby("Source File", dropna=False)
        .size()
        .reset_index(name="Rows Loaded")
        .sort_values("Rows Loaded", ascending=False)
    )
    st.dataframe(file_summary, use_container_width=True, hide_index=True)

render_section("1) Data Mapping", "Map source columns to advisor, classification, metrics, and date.")
all_cols = list(raw_df.columns)
none_opt = ["(none)"] + all_cols

advisor_guess = guess_column(all_cols, ["advisor name", "advisor", "agent name", "agent"])
class_guess = guess_column(all_cols, ["classification", "class", "rank", "tier"])
ac_guess = guess_column(all_cols, ["agency credits", "ac", "annualized", "ape"])
nsc_guess = guess_column(all_cols, ["net sales credits", "nsc", "net sales"])
lives_guess = guess_column(all_cols, ["settled apps", "submitted apps", "lives", "life", "lives count"])
date_guess = guess_column(all_cols, ["process date", "date", "issue date", "submitted date"])

# Standard mapping for your provided row-12 format.
if "Advisor Name" in all_cols:
    advisor_guess = "Advisor Name"
if "Class Code" in all_cols:
    class_guess = "Class Code"
if "Agency Credits" in all_cols:
    ac_guess = "Agency Credits"
if "Net Sales Credits" in all_cols:
    nsc_guess = "Net Sales Credits"
if "Process Date" in all_cols:
    date_guess = "Process Date"
if sheet == "Settled Apps - Details" and "Settled Apps" in all_cols:
    lives_guess = "Settled Apps"
elif sheet == "Submitted Apps - Details" and "Submitted Apps" in all_cols:
    lives_guess = "Submitted Apps"

# If class guess is mostly numeric, prefer Unit when available (usually A/B/C bucket).
if "Unit" in all_cols and series_is_mostly_numeric(raw_df[class_guess]):
    unit_values = raw_df["Unit"].astype(str).str.strip()
    unit_values = unit_values[~unit_values.isin(["", "nan", "None"])]
    if not unit_values.empty and not series_is_mostly_numeric(unit_values):
        class_guess = "Unit"

advisor_col = st.selectbox(
    "Advisor name column", all_cols, index=all_cols.index(advisor_guess), key="advisor_col"
)

# Separate selector because the Google mapping is keyed by *Advisor Code* (not name).
advisor_code_guess = advisor_col
for c in all_cols:
    cl = str(c).strip().lower()
    if "advisor" in cl and "code" in cl:
        advisor_code_guess = c
        break
advisor_code_col = st.selectbox(
    "Advisor code column (used to fix Classification)",  # for the uploaded Dashboard sheet
    all_cols,
    index=all_cols.index(advisor_code_guess) if advisor_code_guess in all_cols else 0,
    key="advisor_code_col",
)

class_col = st.selectbox(
    "Classification column (A/B/C/etc.)",
    all_cols,
    index=all_cols.index(class_guess),
    key="class_col",
)
lives_help = "Use 'Settled Apps' or 'Submitted Apps' for lives."
lives_col = st.selectbox(
    "Lives column (Settled Apps / Submitted Apps)",
    all_cols,
    index=all_cols.index(lives_guess),
    key="lives_col",
    help=lives_help,
)
ac_col = None
nsc_col = None
if sheet != "Submitted Apps - Details":
    ac_index = none_opt.index(ac_guess) if ac_guess in none_opt else 0
    ac_pick = st.selectbox("AC column", none_opt, index=ac_index, key="ac_col")
    nsc_index = none_opt.index(nsc_guess) if nsc_guess in none_opt else 0
    nsc_pick = st.selectbox("NSC column", none_opt, index=nsc_index, key="nsc_col")
    ac_col = None if ac_pick == "(none)" else ac_pick
    nsc_col = None if nsc_pick == "(none)" else nsc_pick
else:
    st.caption("`Submitted Apps - Details` has no AC/NSC columns. AC and NSC are set to 0.")

st.markdown("**Date mapping (choose either Date OR Month+Year)**")
date_col_pick = st.selectbox(
    "Date column", none_opt, index=none_opt.index(date_guess) if date_guess in none_opt else 0
)
date_col = None if date_col_pick == "(none)" else date_col_pick
month_col = None
year_col = None
if not date_col:
    month_col_pick = st.selectbox("Month column", none_opt, index=0)
    year_col_pick = st.selectbox("Year column", none_opt, index=0)
    month_col = None if month_col_pick == "(none)" else month_col_pick
    year_col = None if year_col_pick == "(none)" else year_col_pick
else:
    st.caption("Using Date column only. Month/Year mapping hidden.")
    month_col = None
    year_col = None

st.markdown("**Validation mapping (optional, for Sun Life standards)**")
tenure_guess = guess_column(all_cols, ["tenure", "segment", "advisor tenure"])
coding_guess = guess_column(all_cols, ["coding date", "coded date", "date coded", "process date"])
jfw_guess = guess_column(all_cols, ["jfw"])
start_guess = guess_column(all_cols, ["start"])
pillars_guess = guess_column(all_cols, ["pillar", "4 pillars"])
vul_guess = guess_column(all_cols, ["vul advance", "vul"])
mandatory_guess = guess_column(all_cols, ["mandatory training", "mandatory"])

tenure_pick = st.selectbox(
    "Tenure/segment column", none_opt, index=none_opt.index(tenure_guess) if tenure_guess in none_opt else 0
)
coding_pick = st.selectbox(
    "Coding date column",
    none_opt,
    index=none_opt.index(coding_guess) if coding_guess in none_opt else 0,
    help="Used for Q4-coded Rookie A rule.",
)
jfw_pick = st.selectbox("JFW completed column", none_opt, index=none_opt.index(jfw_guess) if jfw_guess in none_opt else 0)
start_pick = st.selectbox(
    "START completed column", none_opt, index=none_opt.index(start_guess) if start_guess in none_opt else 0
)
pillars_pick = st.selectbox(
    "4 Pillars completed column",
    none_opt,
    index=none_opt.index(pillars_guess) if pillars_guess in none_opt else 0,
)
vul_pick = st.selectbox(
    "VUL Advance completed column", none_opt, index=none_opt.index(vul_guess) if vul_guess in none_opt else 0
)
mandatory_pick = st.selectbox(
    "Mandatory training completed column",
    none_opt,
    index=none_opt.index(mandatory_guess) if mandatory_guess in none_opt else 0,
)

tenure_col = None if tenure_pick == "(none)" else tenure_pick
coding_validation_col = None if coding_pick == "(none)" else coding_pick
jfw_col = None if jfw_pick == "(none)" else jfw_pick
start_col = None if start_pick == "(none)" else start_pick
pillars_col = None if pillars_pick == "(none)" else pillars_pick
vul_col = None if vul_pick == "(none)" else vul_pick
mandatory_col = None if mandatory_pick == "(none)" else mandatory_pick

required_mapping = [advisor_col, class_col, lives_col]
optional_mapping = [c for c in [ac_col, nsc_col] if c]
core_mapping = required_mapping + optional_mapping
if len(set(core_mapping)) < len(core_mapping):
    st.error(
        "Please map different columns for Advisor, Classification, Lives, and optional AC/NSC. "
        "Right now some mappings are duplicated."
    )
    st.stop()

if not date_col and not (month_col and year_col):
    st.warning("Please choose either a Date column or both Month and Year columns.")
    st.stop()

with st.expander("Preview mapped columns"):
    preview_cols = [advisor_col, class_col, lives_col]
    if ac_col:
        preview_cols.append(ac_col)
    if nsc_col:
        preview_cols.append(nsc_col)
    if tenure_col:
        preview_cols.append(tenure_col)
    if coding_validation_col:
        preview_cols.append(coding_validation_col)
    if date_col:
        preview_cols.append(date_col)
    elif month_col and year_col:
        preview_cols.extend([month_col, year_col])
    # Avoid duplicate columns in preview (e.g., Process Date chosen twice).
    preview_cols = list(dict.fromkeys(preview_cols))
    st.dataframe(raw_df[preview_cols].head(10), use_container_width=True)

df = raw_df.copy()
df["Advisor"] = df[advisor_col].astype(str).str.strip()
df["Classification"] = df[class_col].astype(str).str.strip().str.upper()
if "Unit" in df.columns:
    df["Unit"] = df["Unit"].astype(str).str.strip()
else:
    df["Unit"] = "Unknown"

corrected_dashboard_bytes: Optional[bytes] = None
corrected_dashboard_file_name: Optional[str] = None

apply_2026_fix = st.checkbox(
    "Auto-fix Classification using 2026 Advisor Code mapping (Dashboard)",
    value=True,
    help="Matches Advisor Code in your Excel to the 2026 published Google Sheet and updates only matched Classification values (preserves formatting).",
)

if apply_2026_fix:
    with st.spinner("Fetching 2026 Advisor Code → Classification mapping..."):
        code_to_class, target_code_len = fetch_2026_advisor_code_to_classification()

    if not code_to_class:
        st.warning("Could not build 2026 Advisor Code mapping (no rows parsed).")
    else:
        mapped_codes = normalize_advisor_code_series(df[advisor_code_col], target_code_len)
        mapped_cls = mapped_codes.map(code_to_class)
        df["Classification"] = df["Classification"].where(mapped_cls.isna(), mapped_cls)

        # Also correct the uploaded workbook (preserving formatting) for the requested download.
        target_sheet_name = "Dashboard" if "Dashboard" in excel.sheet_names else sheet
        primary_name, primary_bytes = file_payloads[0]
        primary_kind = detect_excel_kind(primary_bytes)
        if primary_kind == "xlsx":
            with st.spinner(
                f"Updating '{target_sheet_name}' sheet classifications in uploaded workbook..."
            ):
                corrected_dashboard_bytes, updated_count = correct_classifications_in_workbook(
                    primary_bytes,
                    sheet_name=target_sheet_name,
                    header_row=int(header_row),
                    advisor_code_header=advisor_code_col,
                    classification_header=class_col,
                    code_to_class=code_to_class,
                    target_code_len=target_code_len,
                    excel_password=excel_password or None,
                )
            corrected_dashboard_file_name = f"corrected_{primary_name}"
            st.caption(
                f"2026 mapping updated {updated_count} Dashboard row(s) in the corrected download."
            )
        else:
            st.warning(
                f"Uploaded file is {primary_kind}; Dashboard formatting-preserving edit "
                "requires a .xlsx (ZIP) file. The in-app classifications are still corrected."
            )

df["AC"] = normalize_number(df[ac_col]) if ac_col else 0
df["NSC"] = normalize_number(df[nsc_col]) if nsc_col else 0
df["Lives"] = normalize_number(df[lives_col])
df["Tenure Raw"] = df[tenure_col].astype(str).str.strip() if tenure_col else ""
if coding_validation_col:
    coding_dt = pd.to_datetime(df[coding_validation_col], errors="coerce")
elif date_col:
    coding_dt = pd.to_datetime(df[date_col], errors="coerce")
else:
    coding_dt = pd.Series(pd.NaT, index=df.index, dtype="datetime64[ns]")
df["Coding Quarter"] = coding_dt.dt.quarter.fillna(0).astype(int)
df["JFW Done"] = normalize_flag(df[jfw_col]) if jfw_col else False
df["START Done"] = normalize_flag(df[start_col]) if start_col else False
df["Pillars Done"] = normalize_flag(df[pillars_col]) if pillars_col else False
df["VUL Advance Done"] = normalize_flag(df[vul_col]) if vul_col else False
df["Mandatory Training Done"] = normalize_flag(df[mandatory_col]) if mandatory_col else False
if sheet != "Submitted Apps - Details" and (not ac_col or not nsc_col):
    st.info("AC/NSC not available in this sheet; missing values are treated as 0.")
df = build_period_columns(df, date_col, month_col, year_col)
df = df[
    ~df["Advisor"].str.lower().str.contains("total", na=False)
    & ~df["Classification"].str.lower().str.contains("total", na=False)
].copy()

if series_is_mostly_numeric(df["Classification"]):
    st.info(
        "Classification values look numeric codes. If you want A/B/C groups, choose `Unit` "
        "or provide your class-code-to-A/B/C mapping rules."
    )

render_section("2) Analysis Filters", "Narrow down classification and time periods for focused insights.")
classes = sorted({str(c) for c in df["Classification"].dropna() if str(c).strip()})
units = sorted({str(u) for u in df["Unit"].dropna() if str(u).strip()})
if not units:
    units = ["Unknown"]
period_months = sort_period_labels(
    list({str(p) for p in df["Period Month"].fillna("Unknown")}), freq="M"
)
period_quarters = sort_period_labels(
    list({str(p) for p in df["Period Quarter"].fillna("Unknown")}), freq="Q"
)
period_years = sorted(
    [y for y in {str(p) for p in df["Period Year"].fillna("Unknown")} if y != "Unknown"]
)
if "Unknown" in set(df["Period Year"].fillna("Unknown").astype(str)):
    period_years.append("Unknown")

st.sidebar.title("Dashboard Filters")
with st.sidebar.expander("How to use this dashboard", expanded=False):
    st.markdown(
        "- Upload one or more files with the same layout.\n"
        "- Pick the target detail sheet and map columns once.\n"
        "- Use filters to focus the story by period/classification.\n"
        "- Download full Excel report with charts and summaries."
    )

with st.sidebar.container(border=True):
    st.markdown('<div class="sl-filter-box-title">Classification</div>', unsafe_allow_html=True)
    selected_classes = st.multiselect("Classification", classes, default=classes, label_visibility="collapsed")
    st.caption("Unit")
    selected_units = st.multiselect("Unit", units, default=units, label_visibility="collapsed")
    search_term = st.text_input(
        "Search advisor or classification",
        value="",
        placeholder="Type advisor name or class...",
        help="Quickly filter dashboard records by advisor name or classification.",
    )
# Build year<->quarter<->month links from available data.
period_pairs = (
    df[["Period Year", "Period Quarter", "Period Month"]]
    .dropna()
    .astype(str)
    .drop_duplicates()
)
year_to_quarters: dict[str, list[str]] = {}
quarter_to_year: dict[str, str] = {}
month_to_quarter: dict[str, str] = {}
quarter_to_months: dict[str, list[str]] = {}
month_to_year: dict[str, str] = {}
for _, row in period_pairs.iterrows():
    year_label = str(row["Period Year"])
    month_label = str(row["Period Month"])
    quarter_label = str(row["Period Quarter"])
    if (
        year_label
        and year_label != "Unknown"
        and month_label
        and month_label != "Unknown"
        and quarter_label
        and quarter_label != "Unknown"
    ):
        year_to_quarters.setdefault(year_label, [])
        if quarter_label not in year_to_quarters[year_label]:
            year_to_quarters[year_label].append(quarter_label)
        quarter_to_year[quarter_label] = year_label
        month_to_quarter[month_label] = quarter_label
        month_to_year[month_label] = year_label
        quarter_to_months.setdefault(quarter_label, []).append(month_label)

saved_months = st.session_state.get("filter_months", period_months)
saved_quarters = st.session_state.get("filter_quarters", period_quarters)
saved_years = st.session_state.get("filter_years", period_years)
saved_months = [m for m in saved_months if m in period_months] or period_months
saved_quarters = [q for q in saved_quarters if q in period_quarters] or period_quarters
saved_years = [y for y in saved_years if y in period_years] or period_years

month_options = period_months
quarter_options = period_quarters
year_options = period_years

# If a subset of years is selected, quarters are limited to those year(s).
if 0 < len(saved_years) < len(period_years):
    allowed_quarters = {
        quarter
        for year in saved_years
        for quarter in year_to_quarters.get(year, [])
    }
    constrained_quarters = [q for q in period_quarters if q in allowed_quarters]
    if constrained_quarters:
        quarter_options = constrained_quarters
        saved_quarters = [q for q in saved_quarters if q in quarter_options] or quarter_options

# If a subset of quarters is selected, months are limited to those quarter(s).
if 0 < len(saved_quarters) < len(quarter_options):
    allowed_months = {
        month
        for quarter in saved_quarters
        for month in quarter_to_months.get(quarter, [])
    }
    constrained_months = [m for m in month_options if m in allowed_months]
    if constrained_months:
        month_options = constrained_months
        saved_months = [m for m in saved_months if m in month_options] or month_options

# If a subset of months is selected, quarters are limited to those month(s).
if 0 < len(saved_months) < len(month_options):
    allowed_quarters = {month_to_quarter.get(month) for month in saved_months}
    constrained_quarters = [
        q for q in quarter_options if q in allowed_quarters and q is not None
    ]
    if constrained_quarters:
        quarter_options = constrained_quarters
        saved_quarters = [q for q in saved_quarters if q in quarter_options] or quarter_options

# If a subset of quarters is selected, years are limited to those quarter(s).
if 0 < len(saved_quarters) < len(quarter_options):
    allowed_years = {quarter_to_year.get(quarter) for quarter in saved_quarters}
    constrained_years = [y for y in period_years if y in allowed_years and y is not None]
    if constrained_years:
        year_options = constrained_years
        saved_years = [y for y in saved_years if y in year_options] or year_options

with st.sidebar.container(border=True):
    st.markdown('<div class="sl-filter-box-title">Time Period</div>', unsafe_allow_html=True)
    st.caption("Years")
    selected_years = st.multiselect(
        "Years", year_options, default=saved_years, key="filter_years", label_visibility="collapsed"
    )
    st.caption("Quarters")
    selected_quarters = st.multiselect(
        "Quarters", quarter_options, default=saved_quarters, key="filter_quarters", label_visibility="collapsed"
    )
    st.caption("Months")
    selected_months = st.multiselect(
        "Months", month_options, default=saved_months, key="filter_months", label_visibility="collapsed"
    )

with st.sidebar.container(border=True):
    st.markdown('<div class="sl-filter-box-title">Ranking</div>', unsafe_allow_html=True)
    top_n = st.slider("Top advisors to show", min_value=5, max_value=30, value=10, step=1)
    ranking_metric = st.selectbox("Advisor ranking metric", ["AC", "NSC", "Lives"], index=0)

with st.sidebar.container(border=True):
    st.markdown('<div class="sl-filter-box-title">Performance Targets</div>', unsafe_allow_html=True)
    target_ac = st.number_input("Target AC", min_value=0.0, value=0.0, step=1000.0)
    target_nsc = st.number_input("Target NSC", min_value=0.0, value=0.0, step=1000.0)
    target_lives = st.number_input("Target Lives", min_value=0.0, value=0.0, step=10.0)

filtered = df[
    df["Classification"].isin(selected_classes)
    & df["Unit"].isin(selected_units)
    & df["Period Year"].isin(selected_years)
    & df["Period Month"].isin(selected_months)
    & df["Period Quarter"].isin(selected_quarters)
].copy()

search_text = search_term.strip().lower()
if search_text:
    advisor_match = filtered["Advisor"].astype(str).str.lower().str.contains(search_text, na=False)
    class_match = filtered["Classification"].astype(str).str.lower().str.contains(search_text, na=False)
    filtered = filtered[advisor_match | class_match].copy()

if filtered.empty:
    st.warning("No rows match your filters.")
    st.stop()

render_section("3) Results & Story", "Executive summary, drivers, trends, data quality, and exports.")

monthly_summary = (
    filtered.groupby(["Period Month", "Classification"], dropna=False)[["AC", "NSC", "Lives"]]
    .sum()
    .reset_index()
    .sort_values(["Period Month", "Classification"])
)

quarterly_summary = (
    filtered.groupby(["Period Quarter", "Classification"], dropna=False)[["AC", "NSC", "Lives"]]
    .sum()
    .reset_index()
    .sort_values(["Period Quarter", "Classification"])
)

classification_summary = (
    filtered.groupby("Classification", dropna=False)[["AC", "NSC", "Lives"]]
    .sum()
    .reset_index()
    .sort_values("AC", ascending=False)
)

advisor_detail = (
    filtered.groupby(
        ["Period Month", "Period Quarter", "Classification", "Advisor"], dropna=False
    )[["AC", "NSC", "Lives"]]
    .sum()
    .reset_index()
    .sort_values(["Period Month", "Classification", "Advisor"])
)

validation_base = filtered.copy()
validation_base["AC+NSC"] = validation_base["AC"] + validation_base["NSC"]
advisor_validation = (
    validation_base.groupby(["Advisor", "Classification"], dropna=False)
    .agg(
        {
            "Tenure Raw": "first",
            "Coding Quarter": "max",
            "AC": "sum",
            "NSC": "sum",
            "AC+NSC": "sum",
            "Lives": "sum",
            "JFW Done": "max",
            "START Done": "max",
            "Pillars Done": "max",
            "VUL Advance Done": "max",
            "Mandatory Training Done": "max",
        }
    )
    .reset_index()
)
validation_cols = advisor_validation.apply(evaluate_sunlife_validation, axis=1, result_type="expand")
advisor_validation = pd.concat([advisor_validation, validation_cols], axis=1).sort_values(
    ["Validation Status", "AC+NSC", "Advisor"], ascending=[True, False, True]
)
advisor_validation["Validation Row Key"] = (
    advisor_validation["Advisor"].astype(str).str.strip()
    + "||"
    + advisor_validation["Classification"].astype(str).str.strip()
)

validation_overrides: dict[str, dict[str, bool]] = st.session_state.get("validation_overrides", {})
if validation_overrides:
    for idx, row in advisor_validation.iterrows():
        row_key = str(row["Validation Row Key"])
        row_override = validation_overrides.get(row_key)
        if not row_override:
            continue
        for col in EDITABLE_VALIDATION_COLUMNS:
            if col in row_override:
                advisor_validation.at[idx, col] = bool(row_override[col])
advisor_validation = enforce_training_requirements(advisor_validation)
validation_cols = advisor_validation.apply(evaluate_sunlife_validation, axis=1, result_type="expand")
advisor_validation = pd.concat(
    [advisor_validation.drop(columns=validation_cols.columns, errors="ignore"), validation_cols], axis=1
).sort_values(["Validation Status", "AC+NSC", "Advisor"], ascending=[True, False, True])
advisor_validation["AC Remaining Balance"] = (float(target_ac) - advisor_validation["AC"]).clip(lower=0)
advisor_validation["NSC Remaining Balance"] = (float(target_nsc) - advisor_validation["NSC"]).clip(lower=0)

render_section("Story Dashboard", "Executive summary, validation snapshot, drivers, trends, and details.")
total_ac = filtered["AC"].sum()
total_nsc = filtered["NSC"].sum()
total_lives = filtered["Lives"].sum()
active_advisors = filtered["Advisor"].nunique()
record_count = len(filtered)
quality_invalid_dates = int((filtered["Period Month"] == "Unknown").sum())
quality_missing_advisor = int(filtered["Advisor"].astype(str).str.strip().isin(["", "nan", "None"]).sum())
quality_missing_class = int(
    filtered["Classification"].astype(str).str.strip().isin(["", "nan", "None"]).sum()
)
quality_negative_values = int(((filtered["AC"] < 0) | (filtered["NSC"] < 0) | (filtered["Lives"] < 0)).sum())

monthly_kpi = (
    filtered.groupby("Period Month", dropna=False)[["AC", "NSC", "Lives"]]
    .sum()
    .reset_index()
)
monthly_kpi["SortMonth"] = pd.PeriodIndex(
    monthly_kpi["Period Month"].where(monthly_kpi["Period Month"] != "Unknown", "1900-01"),
    freq="M",
)
monthly_kpi = monthly_kpi.sort_values("SortMonth").drop(columns=["SortMonth"])

known_monthly = monthly_kpi[monthly_kpi["Period Month"] != "Unknown"].copy()
latest_month = None
previous_month = None
if len(known_monthly) >= 1:
    latest_month = known_monthly.iloc[-1]
if len(known_monthly) >= 2:
    previous_month = known_monthly.iloc[-2]

delta_ac = None
delta_nsc = None
delta_lives = None
if latest_month is not None and previous_month is not None:
    delta_ac = latest_month["AC"] - previous_month["AC"]
    delta_nsc = latest_month["NSC"] - previous_month["NSC"]
    delta_lives = latest_month["Lives"] - previous_month["Lives"]

advisor_record_monthly = (
    filtered[filtered["Period Month"] != "Unknown"]
    .groupby("Period Month", dropna=False)
    .agg(
        {
            "Advisor": pd.Series.nunique,
            "Period Month": "size",
        }
    )
    .rename(columns={"Advisor": "Active Advisors", "Period Month": "Records"})
    .reset_index()
)
if not advisor_record_monthly.empty:
    advisor_record_monthly["SortMonth"] = pd.PeriodIndex(
        advisor_record_monthly["Period Month"], freq="M"
    )
    advisor_record_monthly = advisor_record_monthly.sort_values("SortMonth").drop(
        columns=["SortMonth"]
    )

delta_active_advisors = None
delta_records = None
if len(advisor_record_monthly) >= 2:
    latest_counts = advisor_record_monthly.iloc[-1]
    previous_counts = advisor_record_monthly.iloc[-2]
    delta_active_advisors = int(
        latest_counts["Active Advisors"] - previous_counts["Active Advisors"]
    )
    delta_records = int(latest_counts["Records"] - previous_counts["Records"])

kpi1, kpi2, kpi3, kpi4, kpi5 = st.columns(5)
kpi1.metric("Total AC", format_compact(total_ac), "" if delta_ac is None else format_compact(delta_ac))
kpi2.metric("Total NSC", format_compact(total_nsc), "" if delta_nsc is None else format_compact(delta_nsc))
kpi3.metric(
    "Total Lives",
    format_compact(total_lives),
    "" if delta_lives is None else format_compact(delta_lives),
)
kpi4.metric(
    "Active Advisors",
    f"{active_advisors:,}",
    "" if delta_active_advisors is None else f"{delta_active_advisors:+,}",
)
kpi5.metric("Records", f"{record_count:,}", "" if delta_records is None else f"{delta_records:+,}")

status_ac = performance_status(total_ac, target_ac)
status_nsc = performance_status(total_nsc, target_nsc)
status_lives = performance_status(total_lives, target_lives)


def build_status_row(metric: str, actual: float, target: float, status: str) -> dict[str, object]:
    if target <= 0:
        return {
            "Metric": metric,
            "Actual": actual,
            "Target": target,
            "Achievement %": "",
            "Gap to Target": "",
            "Status": status,
            "What it means": "Add a target to evaluate this metric.",
        }
    achievement_pct = (actual / target) * 100
    gap = actual - target
    if achievement_pct >= 100:
        meaning = "Met or exceeded target."
    elif achievement_pct >= 80:
        meaning = "Near target; improve to close the gap."
    else:
        meaning = "Far from target; needs focused action."
    return {
        "Metric": metric,
        "Actual": actual,
        "Target": target,
        "Achievement %": f"{achievement_pct:.1f}%",
        "Gap to Target": gap,
        "Status": status,
        "What it means": meaning,
    }


status_df = pd.DataFrame(
    [
        build_status_row("AC", total_ac, target_ac, status_ac),
        build_status_row("NSC", total_nsc, target_nsc, status_nsc),
        build_status_row("Lives", total_lives, target_lives, status_lives),
    ]
)

classification_ac_chart = (
    monthly_summary.pivot(index="Period Month", columns="Classification", values="AC")
    .fillna(0)
    .reindex([m for m in period_months if m in set(monthly_summary["Period Month"])])
)

top_advisors = (
    advisor_detail.groupby("Advisor", dropna=False)[["AC", "NSC", "Lives"]]
    .sum()
    .reset_index()
    .sort_values(ranking_metric, ascending=False)
    .head(top_n)
)

top_class = classification_summary.iloc[0] if not classification_summary.empty else None
top_advisor_row = top_advisors.iloc[0] if not top_advisors.empty else None
coverage = (top_class["AC"] / total_ac * 100) if top_class is not None and total_ac else 0

if latest_month is not None:
    st.markdown(
        f"In **{latest_month['Period Month']}**, production reached **AC {latest_month['AC']:,.2f}**, "
        f"**NSC {latest_month['NSC']:,.2f}**, and **Lives {latest_month['Lives']:,.0f}**."
    )
    st.markdown(
        f"""
        <div class="sl-insight-hero">
            <div class="sl-insight-kicker">Monthly Production Snapshot</div>
            <div class="sl-insight-title">{latest_month['Period Month']}</div>
            <div class="sl-insight-metrics">
                <div class="sl-insight-metric">
                    <span class="sl-insight-label">AC</span>
                    <span class="sl-insight-value">{format_compact(float(latest_month['AC']))}</span>
                </div>
                <div class="sl-insight-metric">
                    <span class="sl-insight-label">NSC</span>
                    <span class="sl-insight-value">{format_compact(float(latest_month['NSC']))}</span>
                </div>
                <div class="sl-insight-metric">
                    <span class="sl-insight-label">Lives</span>
                    <span class="sl-insight-value">{int(latest_month['Lives']):,}</span>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

insight_lines = []
if top_class is not None:
    insight_lines.append(
        f"- Top classification: **{top_class['Classification']}** contributed "
        f"**AC {top_class['AC']:,.2f}** ({coverage:.1f}% of total AC)."
    )
if top_advisor_row is not None:
    insight_lines.append(
        f"- Leading advisor by {ranking_metric}: **{top_advisor_row['Advisor']}** "
        f"with **{ranking_metric} {top_advisor_row[ranking_metric]:,.2f}**."
    )
if previous_month is not None and latest_month is not None:
    trend_word = "up" if (delta_ac is not None and delta_ac >= 0) else "down"
    insight_lines.append(
        f"- Month-over-month AC is **{trend_word} {abs(delta_ac):,.2f}** "
        f"from {previous_month['Period Month']} to {latest_month['Period Month']}."
    )
if insight_lines:
    with st.container(border=True):
        st.markdown('<div class="sl-summary-block-title">Executive Insights</div>', unsafe_allow_html=True)
        st.markdown("\n".join(insight_lines))

validation_pass_count = int((advisor_validation["Validation Status"] == "Pass").sum())
validation_total = len(advisor_validation)
vna_count = int((advisor_validation["VNA Eligible"] == "Yes").sum())
vmp_count = int((advisor_validation["VMP Eligible"] == "Yes").sum())
sm_eligible_count = int((advisor_validation["SM Appointment Eligible"] == "Yes").sum())
with st.container(border=True):
    st.markdown('<div class="sl-summary-block-title">Validation Snapshot</div>', unsafe_allow_html=True)
    st.markdown(
        f"- Passed standards: **{validation_pass_count}/{validation_total}** advisors\n"
        f"- VNA eligible: **{vna_count}** | VMP eligible: **{vmp_count}** | SM-appointment eligible: **{sm_eligible_count}**"
    )
    st.markdown(
        f"""
        <div class="sl-mini-stat-grid">
            <div class="sl-mini-stat"><span class="k">Passed Standards</span><span class="v">{validation_pass_count}/{validation_total}</span></div>
            <div class="sl-mini-stat"><span class="k">VNA Eligible</span><span class="v">{vna_count}</span></div>
            <div class="sl-mini-stat"><span class="k">VMP Eligible</span><span class="v">{vmp_count}</span></div>
            <div class="sl-mini-stat"><span class="k">SM Eligible</span><span class="v">{sm_eligible_count}</span></div>
        </div>
        """,
        unsafe_allow_html=True,
    )

story_tab, drivers_tab, trends_tab, details_tab = st.tabs(["Story", "Drivers", "Trends", "Details"])
with story_tab:
    st.markdown("**Performance status vs target**")
    st.caption(
        "Status guide: On track = target achieved; Needs attention = at least 80% of target; "
        "Off track = below 80% of target."
    )
    st.dataframe(status_df, use_container_width=True, hide_index=True)
    st.markdown("**Performance composition by classification**")
    comp_col1, comp_col2 = st.columns(2)
    with comp_col1:
        render_chart_card(
            "Classification composition",
            lambda: _render_classification_composition_chart(classification_summary),
        )
    with comp_col2:
        render_chart_card(
            f"Top {top_n} advisors by {ranking_metric}",
            lambda: _render_top_advisors_chart(top_advisors, ranking_metric=ranking_metric),
        )

with drivers_tab:
    st.markdown("**Who drives production?**")
    left_col, right_col = st.columns(2)
    with left_col:
        class_share = classification_summary.copy()
        class_share["AC Share %"] = (
            class_share["AC"] / class_share["AC"].sum() * 100 if class_share["AC"].sum() else 0
        )
        st.dataframe(class_share, use_container_width=True, height=340)
    with right_col:
        advisor_rank = (
            advisor_detail.groupby(["Classification", "Advisor"], dropna=False)[["AC", "NSC", "Lives"]]
            .sum()
            .reset_index()
            .sort_values(ranking_metric, ascending=False)
            .head(top_n)
        )
        st.dataframe(advisor_rank, use_container_width=True, height=340)

with trends_tab:
    trend_col1, trend_col2 = st.columns(2)
    with trend_col1:
        render_chart_card(
            "Monthly momentum (AC / NSC / Lives)",
            lambda: _render_monthly_momentum_chart(monthly_kpi),
        )
    with trend_col2:
        render_chart_card(
            "Monthly AC by classification",
            lambda: _render_monthly_ac_by_classification(monthly_summary, ordered_months=period_months),
        )
    render_chart_card(
        "Quarterly AC trend by classification",
        lambda: (
            st.altair_chart(
                (
                    alt.Chart(
                        quarterly_summary.assign(
                            **{"Period Quarter": quarterly_summary["Period Quarter"].astype(str)}
                        )
                    )
                    .mark_area(opacity=0.55, interpolate="monotone")
                    .encode(
                        x=alt.X(
                            "Period Quarter:N",
                            sort=sort_period_labels(
                                list(
                                    {
                                        str(p)
                                        for p in quarterly_summary["Period Quarter"].fillna("Unknown")
                                    }
                                ),
                                freq="Q",
                            ),
                            title=None,
                            axis=alt.Axis(labelAngle=0),
                        ),
                        y=alt.Y("AC:Q", stack=True, title=None),
                        color=alt.Color("Classification:N", title=None),
                        tooltip=[
                            alt.Tooltip("Period Quarter:N", title="Quarter"),
                            alt.Tooltip("Classification:N"),
                            alt.Tooltip("AC:Q", format=",.2f"),
                        ],
                    )
                    .properties(height=280)
                ),
                use_container_width=True,
            )
            if (_altair_enabled() and not quarterly_summary.empty)
            else st.area_chart(
                quarterly_summary.pivot(
                    index="Period Quarter", columns="Classification", values="AC"
                ).fillna(0),
                height=280,
            )
        ),
    )

with details_tab:
    tab1, tab2, tab3, tab4, tab5 = st.tabs(
        ["Monthly Summary", "Quarterly Summary", "Advisor Details", "Validation Results", "Data Quality"]
    )
    with tab1:
        st.dataframe(monthly_summary, use_container_width=True)
    with tab2:
        st.dataframe(quarterly_summary, use_container_width=True)
    with tab3:
        st.dataframe(advisor_detail, use_container_width=True, height=420)
        st.markdown("**Advisor AC trend (monthly)**")
        advisor_monthly_ac = (
            filtered.groupby(["Period Month", "Advisor"], dropna=False)[["AC"]]
            .sum()
            .reset_index()
        )
        advisor_monthly_ac = advisor_monthly_ac[advisor_monthly_ac["Period Month"] != "Unknown"].copy()
        advisor_monthly_ac["AC"] = pd.to_numeric(advisor_monthly_ac["AC"], errors="coerce").fillna(0.0)
        if not advisor_monthly_ac.empty:
            advisor_monthly_ac["MonthSort"] = pd.PeriodIndex(advisor_monthly_ac["Period Month"], freq="M")
            advisor_monthly_ac = advisor_monthly_ac.sort_values(["MonthSort", "Advisor"]).drop(columns=["MonthSort"])
        advisor_options = sorted(
            {str(a) for a in advisor_monthly_ac["Advisor"].dropna() if str(a).strip()}
        )
        default_advisors = (
            top_advisors["Advisor"].astype(str).head(min(5, len(top_advisors))).tolist()
            if not top_advisors.empty
            else advisor_options[:5]
        )
        selected_advisor_trend = st.multiselect(
            "Select advisors for AC trend",
            advisor_options,
            default=[a for a in default_advisors if a in advisor_options],
            key="advisor_ac_trend_selection",
        )

        if selected_advisor_trend and not advisor_monthly_ac.empty:
            trend_plot_df = advisor_monthly_ac[
                advisor_monthly_ac["Advisor"].isin(selected_advisor_trend)
            ].copy()
            trend_plot_df["Period Month"] = trend_plot_df["Period Month"].astype(str)
            month_order = sort_period_labels(
                list({str(m) for m in trend_plot_df["Period Month"].dropna() if str(m).strip()}),
                freq="M",
            )
            if _altair_enabled():
                _altair_theme()
                trend_chart = (
                    alt.Chart(trend_plot_df)
                    .mark_line(point=True)
                    .encode(
                        x=alt.X("Period Month:N", sort=month_order, title=None, axis=alt.Axis(labelAngle=0)),
                        y=alt.Y("AC:Q", title="AC"),
                        color=alt.Color("Advisor:N", title="Advisor"),
                        tooltip=[
                            alt.Tooltip("Period Month:N", title="Month"),
                            alt.Tooltip("Advisor:N"),
                            alt.Tooltip("AC:Q", format=",.2f"),
                        ],
                    )
                    .properties(height=300)
                )
                st.altair_chart(trend_chart, use_container_width=True)
            else:
                line_df = (
                    trend_plot_df.pivot(index="Period Month", columns="Advisor", values="AC")
                    .fillna(0)
                    .reindex(month_order)
                )
                st.line_chart(line_df, height=300, use_container_width=True)
        else:
            st.caption("Select at least one advisor to view AC trend.")
    with tab4:
        st.caption(
            "Edit training checkboxes below. Validation status updates automatically based on your selections."
        )
        st.caption(
            "Note: Mandatory Training is for SM appointment only. "
            "It affects only `SM Appointment Eligible`, not `Validation Status`, `VNA Eligible`, or `VMP Eligible`."
        )
        validation_status_filter = st.selectbox(
            "Show validation status",
            options=["All", "Pass", "Fail"],
            index=0,
            key="validation_status_filter",
        )
        previous_overrides: dict[str, dict[str, bool]] = dict(st.session_state.get("validation_overrides", {}))
        editor_cols = [
            "Advisor",
            "Classification",
            "AC",
            "NSC",
            "AC Remaining Balance",
            "NSC Remaining Balance",
            "Tenure Raw",
            "Coding Quarter",
            "JFW Done",
            "START Done",
            "Pillars Done",
            "VUL Advance Done",
            "Mandatory Training Done",
            "Validation Requirement",
            "Validation Status",
            "Validation Reason",
            "VNA Eligible",
            "VMP Eligible",
            "SM Appointment Eligible",
        ]
        filtered_validation = advisor_validation.copy()
        if validation_status_filter != "All":
            filtered_validation = filtered_validation[
                filtered_validation["Validation Status"] == validation_status_filter
            ].copy()
        editor_source = filtered_validation.set_index("Validation Row Key")[editor_cols].copy()
        edited_validation = st.data_editor(
            editor_source,
            use_container_width=True,
            height=420,
            hide_index=True,
            disabled=[
                "Advisor",
                "Classification",
                "AC",
                "NSC",
                "AC Remaining Balance",
                "NSC Remaining Balance",
                "Tenure Raw",
                "Coding Quarter",
                "Validation Requirement",
                "Validation Status",
                "Validation Reason",
                "VNA Eligible",
                "VMP Eligible",
                "SM Appointment Eligible",
            ],
            column_config={
                "AC": st.column_config.NumberColumn("AC", format="%.2f"),
                "NSC": st.column_config.NumberColumn("NSC", format="%.2f"),
                "AC Remaining Balance": st.column_config.NumberColumn("AC Remaining", format="%.2f"),
                "NSC Remaining Balance": st.column_config.NumberColumn("NSC Remaining", format="%.2f"),
                "JFW Done": st.column_config.CheckboxColumn("JFW"),
                "START Done": st.column_config.CheckboxColumn("START"),
                "Pillars Done": st.column_config.CheckboxColumn("Pillars"),
                "VUL Advance Done": st.column_config.CheckboxColumn("VUL Advance"),
                "Mandatory Training Done": st.column_config.CheckboxColumn("SM Mandatory Training"),
            },
            key="validation_training_editor",
        )
        edited_validation = edited_validation.reset_index().rename(columns={"index": "Validation Row Key"})
        updated_overrides = {
            str(row["Validation Row Key"]): {
                "JFW Done": bool(row["JFW Done"]),
                "START Done": bool(row["START Done"]),
                "Pillars Done": bool(row["Pillars Done"]),
                "VUL Advance Done": bool(row["VUL Advance Done"]),
                "Mandatory Training Done": bool(row["Mandatory Training Done"]),
            }
            for _, row in edited_validation.iterrows()
        }
        st.session_state["validation_overrides"] = updated_overrides
        if updated_overrides != previous_overrides:
            st.rerun()

        advisor_validation = advisor_validation.merge(
            edited_validation[
                [
                    "Validation Row Key",
                    "JFW Done",
                    "START Done",
                    "Pillars Done",
                    "VUL Advance Done",
                    "Mandatory Training Done",
                ]
            ],
            on="Validation Row Key",
            how="left",
            suffixes=("", "_edited"),
        )
        for col in EDITABLE_VALIDATION_COLUMNS:
            advisor_validation[col] = advisor_validation[f"{col}_edited"].fillna(advisor_validation[col]).astype(bool)
        advisor_validation = advisor_validation.drop(
            columns=[f"{col}_edited" for col in EDITABLE_VALIDATION_COLUMNS],
            errors="ignore",
        )
        advisor_validation = enforce_training_requirements(advisor_validation)
        validation_cols = advisor_validation.apply(evaluate_sunlife_validation, axis=1, result_type="expand")
        advisor_validation = pd.concat(
            [advisor_validation.drop(columns=validation_cols.columns, errors="ignore"), validation_cols], axis=1
        ).sort_values(["Validation Status", "AC+NSC", "Advisor"], ascending=[True, False, True])
        advisor_validation["AC Remaining Balance"] = (float(target_ac) - advisor_validation["AC"]).clip(lower=0)
        advisor_validation["NSC Remaining Balance"] = (float(target_nsc) - advisor_validation["NSC"]).clip(lower=0)
    with tab5:
        quality_df = pd.DataFrame(
            [
                {"Check": "Rows with unknown/invalid date", "Count": quality_invalid_dates},
                {"Check": "Rows with missing advisor", "Count": quality_missing_advisor},
                {"Check": "Rows with missing classification", "Count": quality_missing_class},
                {"Check": "Rows with negative AC/NSC/Lives", "Count": quality_negative_values},
            ]
        )
        st.dataframe(quality_df, use_container_width=True, hide_index=True)

quality_df = pd.DataFrame(
    [
        {"Check": "Rows with unknown/invalid date", "Count": quality_invalid_dates},
        {"Check": "Rows with missing advisor", "Count": quality_missing_advisor},
        {"Check": "Rows with missing classification", "Count": quality_missing_class},
        {"Check": "Rows with negative AC/NSC/Lives", "Count": quality_negative_values},
    ]
)
executive_summary = pd.DataFrame(
    [
        {"Metric": "Total AC", "Value": total_ac},
        {"Metric": "Total NSC", "Value": total_nsc},
        {"Metric": "Total Lives", "Value": total_lives},
        {"Metric": "Active Advisors", "Value": active_advisors},
        {"Metric": "Records", "Value": record_count},
    ]
)
export_bytes = to_excel_bytes(
    {
        "Executive Summary": executive_summary,
        "Target Status": status_df,
        "Data Quality": quality_df,
        "Monthly Summary": monthly_summary,
        "Quarterly Summary": quarterly_summary,
        "Advisor Details": advisor_detail,
        "Validation Results": advisor_validation.drop(
            columns=["Validation Row Key", "AC+NSC"], errors="ignore"
        ),
        "Filtered Raw Data": filtered,
    }
)

st.markdown("**Download**")
download_col1, download_col2 = st.columns(2)
with download_col1:
    st.download_button(
        "Download full report (Excel)",
        data=export_bytes,
        file_name="manpower_validation_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
with download_col2:
    if corrected_dashboard_bytes is not None and corrected_dashboard_file_name is not None:
        st.download_button(
            "Download corrected uploaded XLSX (Dashboard)",
            data=corrected_dashboard_bytes,
            file_name=corrected_dashboard_file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    st.download_button(
        "Download advisor details (CSV)",
        data=advisor_detail.to_csv(index=False).encode("utf-8"),
        file_name="advisor_details.csv",
        mime="text/csv",
    )